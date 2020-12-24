from django.shortcuts import render
from django.http import HttpResponse

import pandas as pd
import numpy as np
import os

import openpyxl as opx
from .models import Restoraunts
def table(request):

    tables = [
        'xlsxResults/BurgerKing.xlsx',
        'xlsxResults/KFC_2.xlsx',
        'xlsxResults/McDonalds_2.xlsx'
    ]
    f = Restoraunts.objects.all()
    for s in f:
        s.delete()
    for file in tables:
        wb = opx.load_workbook(file)
        ws = wb.get_sheet_by_name('Worked Sheet')
        sheet = 1
        while ws[f'A{sheet}'].value != ws[f'E{sheet}'].value:
            try:
                Restoraunts.objects.create(name=file.split('/')[-1].split('.xlsx')[0], long=ws[f'B{sheet}'].value, width=ws[f'C{sheet}'].value)
            except:
                pass
            sheet += 1
    return HttpResponse(request)



def index(request):
    rests = Restoraunts.objects.all()
    PD = []
    cBK = cMD = cKFC = 0
    McBK = McMD = McKFC = 0
    for rest in rests:
        if rest.name == 'McDonalds_2':
            cMD += 1
            if (int(round(float(rest.long))) == 55 or int(round(float(rest.long))) == 56 or int(round(float(rest.long))) == 54 )and (int(round(float(rest.width))) == 37 or int(round(float(rest.width))) == 36 or int(round(float(rest.width))) == 38):
                McMD += 1
                PD.append([str(rest.name).split('_')[0], float(rest.long), float(rest.width)])
        elif rest.name == "KFC_2":
            cKFC += 1
            if (int(round(float(rest.long))) == 55 or int(round(float(rest.long))) == 56 or int(round(float(rest.long))) == 54 )and (int(round(float(rest.width))) == 37 or int(round(float(rest.width))) == 36 or int(round(float(rest.width))) == 38):
                McKFC += 1
                PD.append([str(rest.name).split('_')[0], float(rest.long), float(rest.width)])
        elif rest.name == "BurgerKing":
            cBK += 1
            if (int(round(float(rest.long))) == 55 or int(round(float(rest.long))) == 56 or int(round(float(rest.long))) == 54 )and (int(round(float(rest.width))) == 37 or int(round(float(rest.width))) == 36 or int(round(float(rest.width))) == 38):
                McBK += 1
                PD.append([rest.name, float(rest.long), float(rest.width)])

    statics = pd.Series(PD).shift(2)
    s = []
    for i in statics:
        print(i)
    wb = opx.Workbook()
    ws = wb.create_sheet('Data')
    for p in PD:
        ws.append(p)
    else:
        try:
            os.mkdir('statics')
        except:
            pass
        wb.save('statics/Data.xlsx')

    pers ={'BK': round(McBK / (McBK + McKFC + McMD)*100),
    'Mc': round(McMD / (McBK + McKFC + McMD)*100),
    'KFC': round(McKFC / (McBK + McKFC + McMD)*100)}
    return render(request, 'index.html', {'cBK': cBK,'cMD': cMD,'cKFC': cKFC,'McBK': McBK,'McMD': McMD,'McKFC': McKFC,'per':pers })