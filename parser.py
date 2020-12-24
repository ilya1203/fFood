from selenium import webdriver
from bs4 import BeautifulSoup as bs
import requests

import openpyxl as opx
import os

from geopy.geocoders import Nominatim

""""
https://burgerking.ru/restaurants

https://www.kfc.ru/restaurants

https://mcdonalds.ru/api/restaurants/polygons/visible?leftX=30.10655144287108&rightX=30.51853874755857&topY=59.99334316105794&bottomY=59.889990404191316
"""


op = webdriver.ChromeOptions()
op.add_argument('headless')
chromedriver = os.getcwd() + '/chromedriver'


class defFunction:
    def writeXLSX(name='file', args=[]):
        wb = opx.Workbook()
        ws = wb.create_sheet('Worked Sheet')
        for a in args:
            try:
                ws.append(a)
            except:
                pass
        else:
            wb.save(f'{name}.xlsx')
    def cleat(txt):
        r = ''
        nw = ''
        for t in txt.split():
            if t != 'г.' and t != 'д.' and t != 'ул.' and t != 'пер.' and  'просп.' not in t  and t != 'ТРЦ' and t !=  'ТЦ' and t != 'ТК' and t != 'ТД' and t != 'ТРК' and t != 'Имени':
                if 'корп.' in t:
                    break
                try:
                    r = r + t.split('И.М.')[-1]
                except:
                    r = r + t
                r = r + ' '
        else:
            for w in r:

                if w != ',':

                    nw = nw + w.lower()
            else:

                nw = nw.split('«')[0]

                return nw





class parser:

    def bk(U=0):
        URL = 'https://burgerking.ru/restaurant-locations-json-reply-new/'
        HEADERS = {'Accept': 'text/css,*/*;q=0.1',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0'}
        sessoin = requests.Session()
        request = sessoin.get(URL, headers=HEADERS)
        soup = bs(request.content, 'html.parser')

        rest = []
        for i in str(soup).split('}'):
            for j in i.split('{'):
                if j != '' and j != '[' and j != ',' and j != ']':
                    x = j.split('"latitude":"')[-1].split('"')[0]
                    y = j.split('longitude":"')[-1].split('"')[0]
                    rest.append([str(len(rest)+1), x,  y])
                    print(rest[-1])
        else:
            try:
                os.mkdir('xlsxResults')
            except:
                pass
            defFunction.writeXLSX('xlsxResults/BurgerKing', rest)


    def kfc(U=0):
        URL = 'https://www.afisha.ru/network/484/?page='

        HEADERS = {'Accept': 'text/css,*/*;q=0.1',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0'}
        sessoin = requests.Session()
        rest = []
        w = 0
        # global op, chromedriver
        # driver = webdriver.Chrome(chromedriver, options=op)
        for i in range(1, 24):
            request = sessoin.get(URL+str(i), headers=HEADERS)
            soup = bs(request.content, 'html.parser')
            for ad in soup.find_all('div', attrs={'class': 'new-list__item-content'}):
            # driver.get(URL+str(i))
                s = ad.text
                try:
                    s = s.split('адрес')[-1]
                except:
                    pass
                try:
                    s = s.split('метро')[0]
                except:
                    pass
                s = defFunction.cleat(s)
                print(s)
                w += 1
                rest.append([w, s])
        else:
            defFunction.writeXLSX('xlsxResults/KFC', rest)



    def mac(U=0):
        URL = 'http://mcdonalds-lem.ru/goroda/'
        HEADERS = {'Accept': 'text/css,*/*;q=0.1',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0'}
        sessoin = requests.Session()
        request = sessoin.get(URL, headers=HEADERS)
        soup = bs(request.content, 'html.parser')

        rest = []
        city = []
        for i in soup.find_all('li'):
            if '<a' in str(i) and '/goroda/' in str(i) and '..' not in str(i):
                print(str(i).split('href="')[-1].split('"')[0])
                city.append(str(i).split('href="')[-1].split('"')[0])
        w = 1

        geolocator = Nominatim()
        for c in city:
            try:
                request = sessoin.get(c, headers=HEADERS)
                soup = bs(request.content, 'html.parser')
                for ad in   soup.find_all('p'):
                    if "<b>Адрес:</b>" in str(ad):

                        try:

                            r = defFunction.cleat(str(str(ad).split('</b>')[-1].split('</p>')[0]))
                            print(r)
                            # location = geolocator.geocode(r)
                            # print(location)
                            rest.append([w]+[r])
                            print(rest[-1])
                            w += 1
                        except Exception as ex:
                            print(f'{ex}')
            except:
                pass
        else:
            defFunction.writeXLSX('xlsxResults/McDonalds', rest)
            print(len(rest))


class gmaps():


    def get_coords(File=''):
        wb = opx.load_workbook(f'{File}.xlsx')
        ws = wb.get_sheet_by_name('Worked Sheet')
        s = 1
        wb2 = opx.Workbook()
        ws2 = wb2.create_sheet("Worked Sheet")
        geolocator = Nominatim()
        while ws[f"A{s}"].value != ws[f"G{s}"].value:
            address = ws[f'B{s}'].value
            while True:
                if len(address.split()) == 0:
                    break
                try:
                    print(address)
                    location = geolocator.geocode(address)
                    ws2.append([s, str(location[-1]).split(',')[0].split('(')[-1], str(location[-1]).split(',')[-1].split(')')[0]])
                    print(location)
                    break
                except Exception as ex:
                    print(f'Err: {ex}')
                    u = address
                    address = ''
                    for i in range(len(u.split())-1):
                        address = address + ' ' + u.split()[i]
            s += 1
        wb2.save(f"{File}_2.xlsx")




gmaps.get_coords('xlsxResults/McDonalds')